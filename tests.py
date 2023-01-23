import openpyxl
from test_pages.files import test_pages
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

PARSER = "html5lib"


def main():
    with open(test_pages("graderesult.php.html")) as file:
        html = file.read()

        soup = BeautifulSoup(html, PARSER)
        table = soup.select('.ewReportTable')
        df = pd.read_html(str(table))[0]
        df.to_excel("file.xlsx", engine='openpyxl')

        # open the excel file
        workbook = openpyxl.load_workbook("file.xlsx")
        worksheet: Worksheet = workbook.active
        worksheet.delete_rows(1, 1)
        worksheet.delete_cols(1, 1)

        # format the cells
        for row in worksheet.iter_rows():
            for c in row:
                cell: Cell = c
                cell.fill = PatternFill("solid", fgColor="DDDDDD")
                cell.border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal='center',
                                           vertical='center')
                cell.font = Font(name='Arial', size=10)

        workbook.save("file.xlsx")


if __name__ == '__main__':
    print(main())
