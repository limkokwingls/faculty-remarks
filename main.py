import asyncio
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pick import pick
from rich import print
from rich.console import Console
from rich.prompt import Prompt

from browser import Browser
from credentials import read_credentials, write_credentials
from model import BorderlineObject
from utils.common import convert_list_to_dict
from workbook_reader import (
    generate_remarks,
    get_border_line_objects,
    get_remark_col,
    get_student_numbers,
)

console = Console()
error_console = Console(stderr=True, style="bold red")
browser = Browser()


def input_username_and_password():
    display_name = browser.login()
    if not display_name:
        error_console.print("Invalid Credentials")
    return display_name


async def login():
    username, password, logged_in = None, None, False
    logged_in = await browser.login()
    while not logged_in:
        logged_in = input_username_and_password()


async def get_cms_marks(worksheet: Worksheet, progress="1/1)"):
    label = worksheet.title
    student_numbers = list(get_student_numbers(worksheet).values())
    results = {}
    with console.status(f"{progress} Reading {label} transcripts..."):
        tasks = []
        for it in student_numbers:
            tasks.append(asyncio.create_task(browser.read_transcript(it, 2)))
        data = await asyncio.gather(*tasks)

        for i, it in enumerate(data):
            results[student_numbers[i]] = it

    transformed_results = {}
    # change results from CourseGrades to dict
    for std_number, course_grades in results.items():
        transformed_data = []
        for it in course_grades:
            transformed_data.append({it.course.code: it.marks})
        transformed_results[std_number] = transformed_data

    # print(transformed_results)
    # exit()

    return convert_list_to_dict(transformed_results)


def open_file():
    file = None
    while file == None:
        try:
            file_path = Prompt.ask("Excel File:", default="data.xlsx")
            file_path = file_path.strip('"')
            if Path(file_path).is_file():
                file = file_path
        except Exception as e:
            error_console.print("Error:", e)
    return file


async def add_remarks(workbook: Workbook, file):
    for i, ws in enumerate(workbook):
        sheet: Worksheet = ws
        cms_marks = await get_cms_marks(
            sheet, progress=f"{i+1}/{len(workbook.sheetnames)})"
        )
        remarks = generate_remarks(sheet, cms_marks)
        remarks_col = get_remark_col(sheet)
        for it in remarks:
            cell: Cell = sheet.cell(row=it, column=remarks_col)
            cell.value = remarks[it]

    workbook.save(file)


def print_borderlines(workbook: Workbook):
    borderlines = []
    for s in workbook:
        data = get_border_line_objects(s)
        for it in data:
            if it.is_borderline():
                borderlines.append(it)

    return borderlines


async def main():
    while not browser.logged_in:
        await login()

    file = open_file()
    with console.status(f"Loading file..."):
        workbook: Workbook = openpyxl.load_workbook(file)

    _, option = pick(["Add Remarks", "Check Borderlines"])
    if option == 0:
        await add_remarks(workbook, file)
    elif option == 1:
        print(print_borderlines(workbook))

    print("[bold blue]Done!")


if __name__ == "__main__":
    # while True:
    asyncio.run(main())
