from turtle import title
import openpyxl
from test_pages.files import test_pages
from bs4 import BeautifulSoup, ResultSet, Tag
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.dataframe import dataframe_to_rows

from utils.excel import delete_empty_columns, fit_column_width, is_merged_cell, set_value

PARSER = "html5lib"


def format_sheet(sheet: Worksheet):
    first_row = 4
    last_col = sheet.max_column

    sheet.row_dimensions[first_row].height = 120  # type:ignore

    first_cell = sheet['A4']
    first_cell.alignment = Alignment(
        vertical='center', horizontal='left', wrap_text=True)

    for row in sheet.iter_rows(min_row=first_row):
        for c in row:
            cell: Cell = c
            if cell.column == last_col:
                sheet.merge_cells(start_row=cell.row, end_row=cell.row,
                                  start_column=last_col, end_column=last_col+3)
    for row in sheet.iter_rows(min_row=first_row):
        for c in row:
            cell: Cell = c
            cell.border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

    fit_column_width(sheet, 'B')
    fit_column_width(sheet, 'C')


def write_to_sheet(sheet: Worksheet, html_table: ResultSet[Tag]):
    for tr_i, tr in enumerate(html_table, start=1):
        col_i = 1
        for td in tr.find_all('td'):
            cell: Cell = sheet.cell(row=tr_i, column=col_i)
            text = td.get_text(strip=True)
            if tr_i == 4 and col_i == 1:
                text = td.get_text(separator='\n', strip=True).splitlines()
                text = "\n".join(text[:-1])
            if not is_merged_cell(sheet, cell):
                set_value(cell, text)

                if td.get('colspan'):
                    span = int(td.get('colspan'))
                    col_i += span
                    sheet.merge_cells(start_row=cell.row, end_row=cell.row,
                                      start_column=cell.column, end_column=cell.column+(span-1))
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')
                else:
                    col_i += 1
    format_sheet(sheet)


def html_to_excel():

    workbook = Workbook()
    sheet: Worksheet = workbook.active

    with open(test_pages("graderesult.php.html")) as file:
        html = file.read()
        soup = BeautifulSoup(html, PARSER)
        table = soup.select('.ewReportTable tr')
        write_to_sheet(sheet, table)

    workbook.save("data.xlsx")


def main():
    html_to_excel()


if __name__ == '__main__':
    print(main())
