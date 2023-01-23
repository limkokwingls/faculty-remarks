import openpyxl
from test_pages.files import test_pages
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.dataframe import dataframe_to_rows
from utils import is_number
from rich import print


PARSER = "html5lib"


def get_marks_cols(workbook: Workbook):
    courses = {}
    for ws in workbook:
        sheet: Worksheet = ws
        for col in sheet.iter_cols():
            for c in col:
                cell: Cell = c
                if cell.value == 'Mk':
                    col_i = cell.col_idx
                    row_i = cell.row
                    grade_cell: Cell = sheet.cell(row_i-2, col_i)
                    # courses[grade_cell.value] = grade_cell.column
                    courses[grade_cell.column] = grade_cell.value
    return courses


def get_std_column(workbook: Workbook):
    for ws in workbook:
        sheet: Worksheet = ws
        for col in sheet.iter_cols():
            for c in col:
                cell: Cell = c
                if cell.value == 'StudentID':
                    return cell.column
    return 3


def get_remark_col(workbook: Workbook):
    for ws in workbook:
        sheet: Worksheet = ws
        for col in sheet.iter_cols():
            for c in col:
                cell: Cell = c
                if 'remark' in cell.value.lower():
                    return cell.column
    return -1


def get_student_numbers(workbook: Workbook) -> dict[int, str]:
    data = {}
    std_col = get_std_column(workbook)
    for ws in workbook:
        sheet: Worksheet = ws
        for col in sheet.iter_cols():
            for c in col:
                cell: Cell = c
                if cell.column == std_col:
                    if is_number(cell.value):
                        data[cell.row] = cell.value

    return data


def main():

    workbook: Workbook = openpyxl.load_workbook("Results 2022-08.xlsx")

    marks = get_marks_cols(workbook)
    students = get_student_numbers(workbook)

    sheet: Worksheet = workbook.active

    data = {}

    for student_col in students:
        student_number = students[student_col]
        repeat = []
        sup = []
        for mark_col in marks:
            mark_value = sheet.cell(student_col, mark_col).value
            if is_number(mark_value):
                mark_value = int(mark_value)
                if mark_value >= 45 and mark_value < 50:
                    sup.append(marks[mark_col])
                elif mark_value < 45:
                    repeat.append(marks[mark_col])
        remarks = "Proceed"
        if len(repeat) >= 3:
            remarks = "Remain in Semester"

        if len(sup) > 0:
            remarks = f"{remarks}, Sup " + ", ".join(sup)
        if len(repeat) > 0:
            remarks = f"{remarks}, Repeat " + ", ".join(repeat)

        data[student_number] = remarks

    print(data)

    # for ws in workbook:
    #     sheet: Worksheet = ws
    #     for i, row in enumerate(sheet.rows):
    #         cell = row[std_col-1]
    #         print(cell.value)

    # print(std_col)


if __name__ == '__main__':
    print(main())
