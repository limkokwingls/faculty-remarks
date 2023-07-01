import asyncio
from typing import Dict

import openpyxl
import pyperclip
from bs4 import BeautifulSoup, ResultSet, Tag
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from rich.console import Console

import urls
from browser import Browser
from credentials import read_credentials
from html_utils import find_link_in_table, get_background, read_table
from utils.excel import (
    delete_empty_columns,
    fit_column_width,
    is_merged_cell,
    set_value,
)

console = Console()
error_console = Console(stderr=True, style="bold red")
browser = Browser()
PARSER = "html5lib"


semesters = ["02", "04", "06", "08"]
programs = [
    "CBIT",
    "DBIT",
    "DIT",
    # "DMSE",
    # "INT",
    # "BIT",
    # "MSE",
    # "BSCIT",
    # "BSCBIT",
    # "BSCSM",
    # "BSIT",
    # "BSBT",
    # "BSSM",
]


async def download_results() -> dict[str, ResultSet[Tag]]:
    data = {}
    res = await browser.session.get(urls.results_page())
    for program in programs:
        for semester in semesters:
            payload = {
                "course": program,
                "sem": semester,
                "term": "2023-02",
                "country": "LSO",
                "grd": "1",
                "Button1": "View",
            }
            class_name = f"{program}_{semester}"
            with console.status(f"Downloading results for {class_name}..."):
                res = await browser.session.post(urls.results_page(), payload)
                soup = BeautifulSoup(res.text, PARSER)
                table = soup.select(".ewReportTable tr")
                data[class_name] = table
    return data


def format_sheet(sheet: Worksheet):
    first_row = 4
    last_col = sheet.max_column

    sheet.row_dimensions[first_row].height = 120  # type:ignore

    first_cell = sheet["A4"]
    first_cell.alignment = Alignment(
        vertical="center", horizontal="left", wrap_text=True
    )

    for row in sheet.iter_rows(min_row=first_row):
        for c in row:
            cell: Cell = c
            if cell.column == last_col:
                sheet.merge_cells(
                    start_row=cell.row,
                    end_row=cell.row,
                    start_column=last_col,
                    end_column=last_col + 3,
                )
    for row in sheet.iter_rows(min_row=first_row):
        for c in row:
            cell: Cell = c
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

    fit_column_width(sheet, "B")
    fit_column_width(sheet, "C")


def write_to_sheet(sheet: Worksheet, html_table: ResultSet[Tag]):
    for tr_i, tr in enumerate(html_table, start=1):
        col_i = 1
        for td in tr.find_all("td"):
            cell: Cell = sheet.cell(row=tr_i, column=col_i)
            text = td.get_text(strip=True)
            if tr_i == 4 and col_i == 1:
                text = td.get_text(separator="\n", strip=True).splitlines()
                text = "\n".join(text[:-1])
            if not is_merged_cell(sheet, cell):
                set_value(cell, text)
                if td.get("colspan"):
                    span = int(td.get("colspan"))
                    col_i += span
                    sheet.merge_cells(
                        start_row=cell.row,
                        end_row=cell.row,
                        start_column=cell.column,
                        end_column=cell.column + (span - 1),
                    )
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    col_i += 1

                # Add background color if any
                bg_color = openpyxl.styles.PatternFill(  # type: ignore
                    start_color=get_background(td),
                    end_color=get_background(td),
                    fill_type="solid",
                )
                cell.fill = bg_color
    format_sheet(sheet)


async def login():
    username, password = None, None
    credentials = read_credentials()
    if credentials:
        username, password = credentials
        await browser.login(username, password)


async def main():
    while not browser.logged_in:
        await login()

    workbook = Workbook()
    results = await download_results()

    with console.status(f"Writing downloaded data to spreadsheet..."):
        for class_name, table in results.items():
            sheet = workbook.create_sheet(class_name)
            write_to_sheet(sheet, table)

        workbook.save("data.xlsx")


if __name__ == "__main__":
    asyncio.run(main())
